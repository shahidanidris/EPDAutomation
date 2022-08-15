def epdCreator():
    import openpyxl
    import shutil
    import pandas as pd
    from functions import failureData
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    import numpy as np
    from plantSortLogic import plantSort
    import PySimpleGUI as sg
    import os

    # GUI Menu
    sg.theme("DarkTeal2")
    layout = [  [sg.Text("IW69 Excel File: "), sg.Input(), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),), key="-IN-")],
                [sg.Text("List Of Tag Numbers: "), sg.Input(), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),), key="-IN2-")],
                [sg.Text("WO Excel File: "), sg.Input(), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),), key="-IN3-")],
                [sg.Button("Submit")]
                ]
    window = sg.Window('My File Browser', layout, size=(600,150))
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event=="Exit":
            break
        elif event == "Submit":
            sourceFile = values["-IN-"]
            tagExcel = values["-IN2-"]
            woExcel = values["-IN3-"]
            break
    window.close()

    # Read IW69
    tempDF = pd.read_excel(sourceFile)

    firstObsv = input("Enter 1st Observation Date [D/M/YYYY HH:MM AM/PM]: ")
    lastObsv = input("Enter last Observation Date [D/M/YYYY HH:MM AM/PM]: ")

    # Tags-To-Scan Excel reading
    tagExcel_df = pd.read_excel(tagExcel)
    ## Make sure Excel has a column name "tags" with all tag numbers provided under this col
    tagsToScan = tagExcel_df['tags'].tolist()

    # To check number of tag numbers to scan matches
    print("Number of Tag Numbers to Create EPD:", len(tagsToScan))

    # Create a folder named "EPDs"
    if os.path.exists('EPDs'):
        shutil.rmtree('EPDs')
        os.makedirs('EPDs')
    else:
        os.makedirs('EPDs')

    totalEPDcount = 0

    for tagNum in tagsToScan:
        df = failureData(tempDF, tagNum)

        if len(df) == 0:
            continue

        print("Doing EPD for " + str(tagNum) + "...")
        shutil.copy("EPD_TEMPLATE.xlsx", 'filename.xlsx')

        epdFile_wb = openpyxl.load_workbook('filename.xlsx')
        epd_ws = epdFile_wb['EPD']
        failureData_ws = epdFile_wb['FAILURE_DATA']

        failureData_ws['C2'] = df.at[0, 'Description']

        if tagNum.endswith('-TX') and tagNum[3] == 'P':
            failureData_ws['C3'] = 'Pressure TX'
        if tagNum.endswith('-TX') and tagNum[3] == 'F':
            failureData_ws['C3'] = 'Flow TX'
        if tagNum.endswith('-TX') and tagNum[3] == 'L':
            failureData_ws['C3'] = 'Level TX'
        if tagNum.endswith('-TX') and tagNum[3] == 'T':
            failureData_ws['C3'] = 'Temperature TX'
        if tagNum.endswith('-TX') and tagNum[3] == 'Q' and "SILIC" in df.at[0, 'Description']:
            failureData_ws['C3'] = 'SILICA ANALYSER'
        if tagNum.endswith('-TX') and tagNum[3] == 'Q' and "COND" in df.at[0, 'Description']:
            failureData_ws['C3'] = 'CONDUCTIVITY ANALYSER'
        if tagNum.endswith('-CV'):
            failureData_ws['C3'] = 'Control Valve'

        failureData_ws['C4'] = df.at[0, 'Functional Loc.']
        failureData_ws['C13'] = '1st observation date'
        failureData_ws['I13'] = firstObsv

        emptyDF = pd.DataFrame(index=np.arange(len(df)), columns=np.arange(1))
        emptyDF.rename(columns={"0":"TypeOfFailure"})
        emptyDF["TypeOfFailure"] = 'F'
        emptyDF = emptyDF[["TypeOfFailure"]]

        new_df = df[["Notification", "Description.1", "ObjPartCodeText"]]
        new_df = new_df.join(emptyDF)
        new_df = new_df.join(df[["Cause grp. text", "Prob. code text", "Cause code text", "Notif.date", "Order"]])
        new_df = new_df.sort_values(by='Notif.date')
        new_df.reset_index(inplace=True)
        new_df = new_df.drop('index', axis=1)

        wo_df = pd.read_excel(woExcel)
        wo_df = wo_df[['Order', 'Actual Finish']]
        new_df = new_df.merge(wo_df, left_on='Order', right_on='Order', how='inner')
        new_df = new_df.drop('Order', axis=1)

        new_df["TBF"] = 0.0

        TBR_series = new_df["Actual Finish"] - new_df["Notif.date"]
        TBR_series = TBR_series.dt.days
        TBR_series = TBR_series.astype('float64')
        new_df["TBR"] = TBR_series

        rows = dataframe_to_rows(new_df, index=False, header=False)

        for r_idx, row in enumerate(rows, 14):
            for c_idx, value in enumerate(row, 2):
                failureData_ws.cell(row=r_idx, column=c_idx, value=value)

        # E6
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # Date Format
        borderStartRow = 13
        borderEndRow = borderStartRow + len(df) + 2
        while borderStartRow < borderEndRow:
            borderStartCol = 9
            borderEndCol = 11
            while borderStartCol < borderEndCol:
                failureData_ws.cell(row=borderStartRow, column=borderStartCol).number_format = '[$-en-US]m/d/yy h:mm AM/PM;@'
                borderStartCol = borderStartCol + 1
            borderStartRow = borderStartRow + 1

        # MTBF MTBR Format
        borderStartRow = 14
        borderEndRow = borderStartRow + len(df) + 1
        while borderStartRow < borderEndRow:
            borderStartCol = 11
            borderEndCol = 13
            while borderStartCol < borderEndCol:
                failureData_ws.cell(row=borderStartRow, column=borderStartCol).number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                borderStartCol = borderStartCol + 1
            borderStartRow = borderStartRow + 1

        # Put borders
        borderStartRow = 14
        borderEndRow = borderStartRow + len(df) + 1
        while borderStartRow < borderEndRow:
            borderStartCol = 2
            borderEndCol = 13
            while borderStartCol < borderEndCol:
                failureData_ws.cell(row=borderStartRow, column=borderStartCol).border = thin_border
                borderStartCol = borderStartCol + 1
            borderStartRow = borderStartRow + 1

        failureData_ws['C' + str(14 + len(df))] = 'last observation date'
        font = Font(italic=True)
        failureData_ws['C' + str(14 + len(df))].font = font
        failureData_ws['I' + str(14 + len(df))] = lastObsv
        failureData_ws['I' + str(14 + len(df))].font = font

        # Manipulating Excel formula through String Input
        borderStartRow = 14
        borderEndRow = borderStartRow + len(df) + 1
        while borderStartRow < borderEndRow:
            failureData_ws['K' + str(borderStartRow)] = '=(I' + str(borderStartRow) + '-I' + str(borderStartRow - 1) + ')/365'
            borderStartRow = borderStartRow + 1

        # Manipulating Excel formula through String Input (Day Conversion)
        borderStartRow = 14
        borderEndRow = borderStartRow + len(df) + 1
        while borderStartRow < borderEndRow:
            failureData_ws['M' + str(borderStartRow)] = '=K' + str(borderStartRow) + '*365'
            borderStartRow = borderStartRow + 1

        borderStartRow = 14
        failureData_ws['K' + str(14 + len(df) + 1)] = 'General MTBF'
        failureData_ws['L' + str(14 + len(df) + 1)] = '=AVERAGE(K' + str(14) + ':K' + str(borderStartRow + len(df)) + ')'
        failureData_ws['K' + str(14 + len(df) + 2)] = 'General MTTR'
        failureData_ws['L' + str(14 + len(df) + 2)] = '=AVERAGE(L' + str(14) + ':L' + str(borderStartRow + len(df)) + ')'

        sampleSheetName = 'COGEN TRANSMITTER, PRESS TURBIN'
        sheetName = "EPD UK " + plantSort(tagNum) + " " + df.at[0, 'Description']
        
        if "//" in sheetName:
            sheetName = sheetName.replace("//", "-")

        if len(sheetName) > len(sampleSheetName):
            failureData_ws.title = sheetName[0:len(sampleSheetName)]
        else:
            failureData_ws.title = sheetName
        
        epdFilePath = "EPDs/" + sheetName + ".xlsx"
        
        epdFile_wb.save(epdFilePath)

        totalEPDcount = totalEPDcount + 1

    print("Completed EPD files:", totalEPDcount)